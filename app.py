import streamlit as st
import json
import re
import os
import time
from datetime import datetime
import openai
from groq import Groq
import google.generativeai as genai
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from io import BytesIO
import requests
import ast
import tempfile
import subprocess

# ================= PAGE CONFIG =================
st.set_page_config(
    page_title="EOD Report Generator",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ================= HIDE FOOTER =================
st.markdown("""
<style>
    footer {visibility: hidden;}
    .stAppFooter {display: none;}
</style>
""", unsafe_allow_html=True)

# ================= THEME STATE =================
if "theme" not in st.session_state:
    st.session_state.theme = "light"

st.session_state.setdefault("selected_employee_name", "Omkar Patil")
st.session_state.setdefault("selected_employee_position", "Social Media & Digital Marketing Executive")
st.session_state.setdefault("loaded_report", None)
st.session_state.setdefault("loaded_date", None)

def toggle_theme():
    st.session_state.theme = "dark" if st.session_state.theme == "light" else "light"
    st.rerun()

# ================= THEME VARIABLES =================
theme = st.session_state.theme
if theme == "light":
    bg_primary = "#f8f9fa"
    bg_secondary = "#ffffff"
    text_color = "#1a1a1a"
    card_bg = "rgba(255,255,255,0.9)"
    border_color = "rgba(0,0,0,0.08)"
    shadow_color = "rgba(0,0,0,0.08)"
    input_bg = "#ffffff"
    input_text = "#1a1a1a"
    input_border = "#cccccc"
    table_header_bg = "#f0f2f6"
    table_border = "#d0d0d0"
    slot_bg = "rgba(102,126,234,0.06)"
    lunch_bg = "rgba(255,193,7,0.12)"
else:  # dark
    bg_primary = "#0e1117"
    bg_secondary = "#1e1e2a"
    text_color = "#f0f0f0"
    card_bg = "rgba(30,30,50,0.7)"
    border_color = "rgba(255,255,255,0.1)"
    shadow_color = "rgba(0,0,0,0.4)"
    input_bg = "#2a2a3a"
    input_text = "#f0f0f0"
    input_border = "#444466"
    table_header_bg = "#2a2a3e"
    table_border = "#444466"
    slot_bg = "rgba(102,126,234,0.15)"
    lunch_bg = "rgba(255,193,7,0.15)"

header_grad = "linear-gradient(135deg, #667eea 0%, #764ba2 50%, #f093fb 100%)"
preview_bg = bg_secondary
preview_border = border_color

# ================= PROFESSIONAL, THEME-AWARE CSS =================
st.markdown(f"""
<style>
    /* ---------- GLOBAL ---------- */
    * {{
        box-sizing: border-box;
        margin: 0;
        padding: 0;
    }}
    .block-container {{
        padding: 0.3rem 0.6rem 0.2rem 0.6rem !important;
        max-width: 1100px !important;
        margin: 0 auto !important;
    }}
    .stApp {{
        background: {bg_primary};
        color: {text_color} !important;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        transition: background 0.3s ease, color 0.3s ease;
    }}
    /* ---------- HEADER ---------- */
    .main-header {{
        background: {header_grad};
        background-size: 300% 300%;
        animation: gradientFlow 8s ease infinite, fadeInUp 0.8s ease;
        padding: 0.6rem 1.2rem;
        border-radius: 20px;
        color: white;
        text-align: center;
        margin-bottom: 0.8rem;
        margin-top: 1.6rem;
        box-shadow: 0 8px 40px rgba(102, 126, 234, 0.35);
        border: 1px solid rgba(255,255,255,0.2);
        position: relative;
        overflow: hidden;
    }}
    .main-header::after {{
        content: '';
        position: absolute;
        top: -50%;
        left: -50%;
        width: 200%;
        height: 200%;
        background: radial-gradient(circle at 30% 30%, rgba(255,255,255,0.15) 0%, transparent 60%);
        pointer-events: none;
        animation: shine 10s linear infinite;
    }}
    @keyframes fadeInUp {{
        0% {{ opacity: 0; transform: translateY(30px); }}
        100% {{ opacity: 1; transform: translateY(0); }}
    }}
    @keyframes shine {{
        0% {{ transform: translateX(-100%) rotate(20deg); }}
        100% {{ transform: translateX(100%) rotate(20deg); }}
    }}
    @keyframes gradientFlow {{
        0% {{ background-position: 0% 50%; }}
        50% {{ background-position: 100% 50%; }}
        100% {{ background-position: 0% 50%; }}
    }}
    .main-header h1 {{
        font-size: 1.4rem;
        font-weight: 700;
        letter-spacing: -0.5px;
        margin: 0;
        line-height: 1.2;
        text-shadow: 0 2px 20px rgba(0,0,0,0.25);
        position: relative;
        z-index: 2;
        color: white !important;
    }}
    .main-header p {{
        font-size: 0.75rem;
        opacity: 0.9;
        margin: 0.1rem 0 0 0;
        font-weight: 300;
        letter-spacing: 0.3px;
        position: relative;
        z-index: 2;
        color: white !important;
    }}
    /* ---------- ALL TEXT ELEMENTS ---------- */
    .stMarkdown, .stMarkdown p, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, 
    .stMarkdown h4, .stMarkdown h5, .stMarkdown h6, .stMarkdown li, .stMarkdown blockquote,
    .stTextInput label, .stDateInput label, .stSelectbox label, .stSlider label,
    .stCheckbox label, .stRadio label, .stTextArea label {{
        color: {text_color} !important;
    }}
    .stCaption {{
        color: {text_color} !important;
        opacity: 0.7;
    }}
    /* ---------- CARDS ---------- */
    .glass-card {{
        background: {card_bg};
        backdrop-filter: blur(8px);
        border-radius: 12px;
        padding: 0.2rem 0.5rem;
        margin-bottom: 0.2rem;
        box-shadow: 0 4px 20px {shadow_color};
        border: 1px solid {border_color};
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }}
    .glass-card:hover {{
        transform: translateY(-2px);
        box-shadow: 0 8px 30px {shadow_color};
    }}
    /* ---------- SLOT CARDS ---------- */
    .slot-card {{
        background: {slot_bg};
        border-radius: 10px;
        padding: 0.1rem 0.4rem;
        margin-bottom: 0.1rem;
        border: 1px solid {border_color};
        box-shadow: 0 2px 8px {shadow_color};
        transition: all 0.2s ease;
        color: {text_color} !important;
        display: flex;
        align-items: center;
        flex-wrap: wrap;
        gap: 4px;
    }}
    .slot-card:hover {{
        transform: scale(1.005);
        box-shadow: 0 4px 16px rgba(102,126,234,0.15);
        border-color: rgba(102,126,234,0.3);
    }}
    .slot-label {{
        font-weight: 600;
        font-size: 0.7rem;
        color: {text_color} !important;
        min-width: 70px;
        margin-right: 4px;
        letter-spacing: 0.2px;
    }}
    .lunch-card {{
        background: {lunch_bg};
        border: 1px solid rgba(255,193,7,0.2);
        color: {text_color} !important;
    }}
    /* ---------- BUTTONS ---------- */
    .stButton button {{
        padding: 0.25rem 0.5rem !important;
        font-size: 0.8rem !important;
        border-radius: 10px !important;
        transition: all 0.3s ease !important;
        background: {header_grad} !important;
        background-size: 200% 200% !important;
        color: white !important;
        border: none !important;
        font-weight: 600 !important;
        box-shadow: 0 2px 12px rgba(102, 126, 234, 0.3) !important;
        min-height: 36px !important;
        width: 100% !important;
        touch-action: manipulation !important;
        letter-spacing: 0.3px;
    }}
    .stButton button:hover {{
        transform: scale(1.02);
        box-shadow: 0 6px 24px rgba(102, 126, 234, 0.5) !important;
        background-position: 100% 50% !important;
    }}
    .stButton button:active {{
        transform: scale(0.97);
    }}
    .stDownloadButton button {{
        padding: 0.25rem 0.5rem !important;
        font-size: 0.8rem !important;
        border-radius: 10px !important;
        transition: all 0.3s ease !important;
        background: #28a745 !important;
        color: white !important;
        border: none !important;
        font-weight: 600 !important;
        box-shadow: 0 2px 12px rgba(40, 167, 69, 0.3) !important;
        min-height: 36px !important;
        width: 100% !important;
        touch-action: manipulation !important;
        letter-spacing: 0.3px;
    }}
    .stDownloadButton button:hover {{
        transform: scale(1.02);
        box-shadow: 0 6px 24px rgba(40, 167, 69, 0.5) !important;
    }}
    .stDownloadButton button:nth-of-type(2) {{
        background: #dc3545 !important;
        box-shadow: 0 2px 12px rgba(220, 53, 69, 0.3) !important;
    }}
    .stDownloadButton button:nth-of-type(2):hover {{
        box-shadow: 0 6px 24px rgba(220, 53, 69, 0.5) !important;
    }}
    /* ---------- SIDEBAR ---------- */
    .css-1d391kg {{
        background: {bg_secondary} !important;
        backdrop-filter: blur(12px);
        color: {text_color} !important;
        padding: 0.15rem 0.25rem !important;
        border-right: 1px solid {border_color} !important;
        box-shadow: 4px 0 30px {shadow_color} !important;
        transition: background 0.3s ease, color 0.3s ease;
    }}
    .css-1d391kg * {{
        font-size: 0.75rem !important;
        color: {text_color} !important;
    }}
    .css-1d391kg .stSelectbox select,
    .css-1d391kg .stTextInput input,
    .css-1d391kg .stDateInput input {{
        background: {input_bg} !important;
        color: {input_text} !important;
        border: 1px solid {input_border} !important;
        border-radius: 8px !important;
        font-size: 0.75rem !important;
        padding: 0.2rem 0.4rem !important;
        height: 32px !important;
        width: 100% !important;
        transition: background 0.3s ease, color 0.3s ease, border 0.3s ease;
    }}
    .css-1d391kg .stSelectbox select:focus,
    .css-1d391kg .stTextInput input:focus,
    .css-1d391kg .stDateInput input:focus {{
        border-color: #6C63FF !important;
        box-shadow: 0 0 0 3px rgba(108, 99, 255, 0.2) !important;
    }}
    .css-1d391kg .stButton button {{
        background: rgba(108, 99, 255, 0.8) !important;
        color: white !important;
        font-size: 0.7rem !important;
        padding: 0.2rem 0.4rem !important;
        height: 30px !important;
        border: 1px solid {input_border} !important;
        border-radius: 8px !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08) !important;
        min-height: 30px !important;
        transition: background 0.3s ease;
    }}
    .css-1d391kg .stButton button:hover {{
        background: rgba(108, 99, 255, 1) !important;
    }}
    /* ---------- MAIN INPUTS ---------- */
    .stTextInput input, .stDateInput input, .stSelectbox select {{
        background: {input_bg} !important;
        color: {input_text} !important;
        border: 1px solid {input_border} !important;
        border-radius: 10px !important;
        font-size: 0.8rem !important;
        padding: 0.25rem 0.5rem !important;
        height: 36px !important;
        width: 100% !important;
        transition: background 0.3s ease, color 0.3s ease, border 0.3s ease;
    }}
    .stTextInput input:focus, .stDateInput input:focus, .stSelectbox select:focus {{
        border-color: #6C63FF !important;
        box-shadow: 0 0 0 4px rgba(108, 99, 255, 0.1) !important;
    }}
    /* ---------- PREVIEW TABLE ---------- */
    .preview-table {{
        width: 100%;
        border-collapse: collapse;
        margin-top: 4px;
        font-size: 10px;
        color: {text_color} !important;
        border-radius: 10px;
        overflow: hidden;
        box-shadow: 0 2px 12px {shadow_color};
    }}
    .preview-table th {{
        background-color: {table_header_bg};
        font-weight: 600;
        border: 1px solid {table_border};
        padding: 3px 6px;
        text-align: left;
        color: {text_color} !important;
    }}
    .preview-table td {{
        border: 1px solid {table_border};
        padding: 3px 6px;
        text-align: left;
        background: {preview_bg};
        color: {text_color} !important;
    }}
    .preview-table tr:hover td {{
        background: rgba(108, 99, 255, 0.04);
    }}
    /* ---------- COMPACT MOBILE ---------- */
    @media (max-width: 768px) {{
        .block-container {{
            padding: 0.15rem 0.2rem 0.1rem 0.2rem !important;
        }}
        .main-header h1 {{ font-size: 1.1rem !important; }}
        .main-header p {{ font-size: 0.6rem !important; }}
        .stButton button, .stDownloadButton button {{
            font-size: 0.8rem !important;
            padding: 0.3rem 0.2rem !important;
            min-height: 40px !important;
        }}
        .slot-label {{
            font-size: 0.65rem !important;
            min-width: 60px !important;
        }}
        .slot-card {{
            padding: 0.1rem 0.3rem !important;
            margin-bottom: 0.08rem !important;
        }}
        .stTextInput input, .stDateInput input, .stSelectbox select {{
            font-size: 16px !important;
            height: 40px !important;
            padding: 0.25rem 0.3rem !important;
        }}
        .css-1d391kg {{
            padding: 0.1rem 0.15rem !important;
            min-width: 200px !important;
        }}
        .css-1d391kg * {{
            font-size: 0.65rem !important;
        }}
        .stColumns {{
            flex-direction: column !important;
        }}
        .preview-table {{
            font-size: 9px !important;
        }}
        .preview-table th, .preview-table td {{
            padding: 2px 4px !important;
        }}
    }}
    @media (max-width: 480px) {{
        .main-header h1 {{ font-size: 0.9rem !important; }}
        .main-header p {{ font-size: 0.55rem !important; }}
        .slot-label {{
            font-size: 0.6rem !important;
            min-width: 50px !important;
        }}
        .preview-table {{
            font-size: 8px !important;
        }}
        .preview-table th, .preview-table td {{
            padding: 1px 3px !important;
        }}
    }}
</style>
""", unsafe_allow_html=True)

# ================= HEADER =================
st.markdown("""
<div class="main-header">
    <h1>✨ EOD Report Generator</h1>
    <p>Intelligent AI‑powered reports – polished, precise, and beautifully presented</p>
</div>
""", unsafe_allow_html=True)

# ================= ROBUST DATE PARSER =================
def parse_date(date_str):
    if not date_str:
        return datetime.now().date()
    match = re.search(r'(\d{2})/(\d{2})/(\d{4})', date_str)
    if match:
        try:
            return datetime.strptime(match.group(0), "%d/%m/%Y").date()
        except ValueError:
            pass
    match = re.search(r'(\d{4}-\d{2}-\d{2})', date_str)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y-%m-%d").date()
        except ValueError:
            pass
    return datetime.now().date()

# ================= DYNAMIC TIME SLOTS =================
def get_time_slots(lunch_hour):
    slots = []
    for h in range(10, 18):
        start = h
        end = h + 1
        if h < 12:
            start_str = f"{h:02d}:00 am"
        elif h == 12:
            start_str = "12:00 pm"
        else:
            start_str = f"{h-12:02d}:00 pm"
        if end < 12:
            end_str = f"{end:02d}:00 am"
        elif end == 12:
            end_str = "12:00 pm"
        else:
            end_str = f"{end-12:02d}:00 pm"
        slots.append(f"{start_str} to {end_str}")
    return slots

def get_time_slots_short(lunch_hour):
    slots = []
    for h in range(10, 18):
        start = h
        end = h + 1
        if h < 12:
            start_str = f"{h:02d}:00"
        elif h == 12:
            start_str = "12:00"
        else:
            start_str = f"{h-12:02d}:00"
        if end < 12:
            end_str = f"{end:02d}:00"
        elif end == 12:
            end_str = "12:00"
        else:
            end_str = f"{end-12:02d}:00"
        slots.append(f"{start_str}-{end_str}")
    return slots

def create_short_to_full_map(lunch_hour):
    short = get_time_slots_short(lunch_hour)
    full = get_time_slots(lunch_hour)
    return dict(zip(short, full))

def get_full_to_short_map(lunch_hour):
    short_to_full = create_short_to_full_map(lunch_hour)
    return {v: k for k, v in short_to_full.items()}

# ================= REBUILD SCHEDULE FOR LUNCH CHANGE =================
def rebuild_schedule_for_lunch(lunch_hour, old_schedule):
    if not old_schedule:
        return None
    new_slots = get_time_slots(lunch_hour)
    hour_map = {}
    for entry in old_schedule:
        slot_label = entry["slot"]
        match = re.match(r'(\d{1,2}):00 (am|pm)', slot_label.split(' to ')[0])
        if match:
            hour = int(match.group(1))
            if match.group(2) == 'pm' and hour != 12:
                hour += 12
            elif match.group(2) == 'am' and hour == 12:
                hour = 0
            if entry["activity"] != "Lunch Break":
                hour_map[hour] = (entry["activity"], entry["description"])
    new_schedule = []
    for slot_label in new_slots:
        match = re.match(r'(\d{1,2}):00 (am|pm)', slot_label.split(' to ')[0])
        if match:
            hour = int(match.group(1))
            if match.group(2) == 'pm' and hour != 12:
                hour += 12
            elif match.group(2) == 'am' and hour == 12:
                hour = 0
            if hour == lunch_hour:
                new_schedule.append({"slot": slot_label, "activity": "Lunch Break", "description": "Lunch Break"})
            elif hour in hour_map:
                act, desc = hour_map[hour]
                new_schedule.append({"slot": slot_label, "activity": act, "description": desc})
            else:
                new_schedule.append({"slot": slot_label, "activity": "No specific task", "description": "No description provided."})
        else:
            new_schedule.append({"slot": slot_label, "activity": "No specific task", "description": "No description provided."})
    return new_schedule

# ================= CONFIG =================
CONFIG_FILE = ".eod_config.json"
HISTORY_FILE = "history.json"
EMPLOYEES_FILE = "employees.json"

DEFAULT_EMPLOYEE = "Omkar Patil"
DEFAULT_POSITION = "Social Media & Digital Marketing Executive"

PROVIDERS = {
    "Groq (Fastest)": {
        "default_model": "llama-3.1-8b-instant",
        "api_key_required": True,
    },
    "OpenAI (ChatGPT)": {
        "default_model": "gpt-4o-mini",
        "api_key_required": True,
    },
    "Google Gemini": {
        "default_model": "gemini-1.5-flash",
        "api_key_required": True,
    },
    "Ollama (local)": {
        "default_model": "phi3",
        "api_key_required": False,
    }
}

# ============= TEMPLATE MANAGEMENT =============
TEMPLATE_DIR = "templates"
os.makedirs(TEMPLATE_DIR, exist_ok=True)

def get_template_list():
    files = [f for f in os.listdir(TEMPLATE_DIR) if f.endswith('.xlsx')]
    return sorted(files)

def load_template_bytes(filename):
    path = os.path.join(TEMPLATE_DIR, filename)
    if os.path.exists(path):
        with open(path, "rb") as f:
            return f.read()
    if os.path.exists(filename):
        with open(filename, "rb") as f:
            return f.read()
    return None

DEFAULT_TEMPLATE_BYTES = None
DEFAULT_TEMPLATE_FILE = "EOD_SAMPLE_FINAL.xlsx"
if os.path.exists(DEFAULT_TEMPLATE_FILE):
    with open(DEFAULT_TEMPLATE_FILE, "rb") as f:
        DEFAULT_TEMPLATE_BYTES = f.read()

# ============= HELPER =============
def generate_filename(date_obj, employee_name, extension):
    date_str = date_obj.strftime("%d-%m-%y")
    first_name = employee_name.split()[0].upper() if employee_name.strip() else "UNKNOWN"
    return f"EOD_{date_str}_{first_name}.{extension}"

# ============= PERSISTENCE HELPERS =============
def load_json_file(filepath, default=None):
    if os.path.exists(filepath):
        try:
            with open(filepath, "r") as f:
                return json.load(f)
        except:
            return default if default is not None else {}
    return default if default is not None else {}

def save_json_file(filepath, data):
    with open(filepath, "w") as f:
        json.dump(data, f, indent=2)

def load_config():
    return load_json_file(CONFIG_FILE, {})

def save_config(provider, model, api_key=None):
    data = {"provider": provider, "model": model}
    if api_key is not None:
        data["api_key"] = api_key
    save_json_file(CONFIG_FILE, data)

def clear_config():
    if os.path.exists(CONFIG_FILE):
        os.remove(CONFIG_FILE)

# ============= EMPLOYEE & HISTORY =============
def load_employees():
    return load_json_file(EMPLOYEES_FILE, [])

def save_employees(employees):
    save_json_file(EMPLOYEES_FILE, employees)

def add_employee(name, position):
    employees = load_employees()
    if not any(e["name"] == name for e in employees):
        employees.append({"name": name, "position": position})
        save_employees(employees)
    return employees

def delete_employee(name):
    employees = load_employees()
    employees = [e for e in employees if e["name"] != name]
    save_employees(employees)
    return employees

def load_history():
    return load_json_file(HISTORY_FILE, [])

def save_history(history):
    save_json_file(HISTORY_FILE, history)

def add_history_entry(employee_name, position, date, schedule):
    history = load_history()
    history.append({
        "timestamp": datetime.now().isoformat(),
        "employee_name": employee_name,
        "position": position,
        "date": date,
        "schedule": schedule
    })
    save_history(history)

def clear_history():
    if os.path.exists(HISTORY_FILE):
        os.remove(HISTORY_FILE)

# ============= JSON PARSER =============
def extract_and_clean_json(raw_text):
    raw_text = re.sub(r'```json\s*', '', raw_text)
    raw_text = re.sub(r'```\s*', '', raw_text)
    start = raw_text.find('{')
    end = raw_text.rfind('}')
    if start != -1 and end != -1:
        json_str = raw_text[start:end+1]
    else:
        start = raw_text.find('[')
        end = raw_text.rfind(']')
        if start != -1 and end != -1:
            json_str = raw_text[start:end+1]
        else:
            raise ValueError("No JSON-like structure found")
    try:
        return json.loads(json_str)
    except:
        pass
    json_str = re.sub(r'([{,])\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*:', r'\1"\2":', json_str)
    json_str = re.sub(r',\s*([}\]])', r'\1', json_str)
    try:
        return json.loads(json_str)
    except:
        pass
    try:
        py_str = json_str.replace('null', 'None').replace('true', 'True').replace('false', 'False')
        return ast.literal_eval(py_str)
    except:
        pass
    raise ValueError("Could not parse JSON")

# ============= AI GENERATION =============
def generate_schedule(user_tasks, employee_name, position, report_date, provider, api_key, model_name, lunch_hour, progress_callback=None):
    if PROVIDERS[provider]["api_key_required"] and not api_key:
        raise ValueError(f"❌ API key for {provider} is missing. Please enter it in the sidebar under Config → API Key.")

    slot_labels = get_time_slots(lunch_hour)
    lunch_label = slot_labels[lunch_hour - 10]
    short_to_full = create_short_to_full_map(lunch_hour)
    full_to_short = get_full_to_short_map(lunch_hour)

    task_mapping = {}
    if user_tasks:
        for line in user_tasks.split('\n'):
            line = line.strip()
            if not line:
                continue
            match = re.match(r'^(\d{2}:\d{2}-\d{2}:\d{2})\s*:\s*(.*)$', line)
            if match:
                short_time, task = match.groups()
                task = task.strip()
                task_mapping[short_time] = task

    def call_ai(prompt_extra=""):
        full_prompt = f"""
You are an assistant that fills an End‑of‑Day work report.

The report has exactly these 8 hourly slots (lunch break is fixed at **{lunch_label}** – you MUST set activity="Lunch Break" and description="Lunch Break" for that slot):

All slots must be filled. Do not leave any slot empty.

The user has provided the following tasks for specific time slots (short slot → task):
{json.dumps(task_mapping, indent=2)}

Instructions:
- For each time slot, you must generate a **short, concise activity title** (2‑5 words) based on the user's provided task.
- Write a **detailed description** (1‑2 sentences) that explains what the user did.
- If the user's task is "-", set both activity and description to "-".
- If a slot has no user task, invent a reasonable activity and description.
- For the lunch slot, always use activity="Lunch Break" and description="Lunch Break".

Return **only** a valid JSON object with:
- "employee_name" (use the provided name)
- "position" (use the provided position)
- "date" (use the provided date)
- "schedule": an array of objects with keys "slot", "activity", "description". Include exactly the above 8 slots in the correct order.

Use double quotes for all keys and string values. No trailing commas. Do not include any text outside the JSON.

{prompt_extra}

Employee: {employee_name}
Position: {position}
Date: {report_date}
"""
        timeout_seconds = 60 if provider == "Ollama (local)" else 30
        if provider == "Groq (Fastest)":
            client = Groq(api_key=api_key)
            model = model_name or PROVIDERS[provider]["default_model"]
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": full_prompt}],
                temperature=0.2,
                max_tokens=600,
                timeout=timeout_seconds
            )
            return response.choices[0].message.content
        elif provider == "OpenAI (ChatGPT)":
            client = openai.OpenAI(api_key=api_key)
            model = model_name or PROVIDERS[provider]["default_model"]
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": full_prompt}],
                temperature=0.2,
                max_tokens=600,
                timeout=timeout_seconds
            )
            return response.choices[0].message.content
        elif provider == "Google Gemini":
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel(model_name or PROVIDERS[provider]["default_model"])
            response = model.generate_content(
                full_prompt,
                generation_config=genai.types.GenerationConfig(temperature=0.2, max_output_tokens=600),
                request_options={"timeout": timeout_seconds}
            )
            return response.text
        elif provider == "Ollama (local)":
            try:
                requests.get("http://localhost:11434", timeout=5)
            except requests.ConnectionError:
                raise RuntimeError("Ollama is not running. Please start Ollama.")
            client = openai.OpenAI(base_url="http://localhost:11434/v1", api_key="dummy")
            model = model_name or PROVIDERS[provider]["default_model"]
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": full_prompt}],
                temperature=0.2,
                max_tokens=600,
                timeout=timeout_seconds
            )
            return response.choices[0].message.content
        else:
            raise ValueError("Unsupported provider")

    raw_response = None
    data = None
    for attempt in range(2):
        if progress_callback:
            progress_callback(30 + attempt * 20, f"Contacting AI ({provider})... attempt {attempt+1}")
        try:
            extra = "IMPORTANT: You MUST include ALL of the following slots exactly as listed: " + ", ".join(slot_labels) if attempt == 1 else ""
            raw = call_ai(extra)
            raw_response = raw
            parsed = extract_and_clean_json(raw)

            if "schedule" not in parsed or not isinstance(parsed["schedule"], list):
                raise ValueError("No 'schedule' array in response")

            ai_schedule = parsed["schedule"]

            # If we have exactly 8 entries, use index-based mapping
            if len(ai_schedule) == len(slot_labels):
                for i, entry in enumerate(ai_schedule):
                    entry["slot"] = slot_labels[i]
                parsed["schedule"] = ai_schedule
                data = parsed
                break
            else:
                # Fallback: match by hour
                hour_to_entry = {}
                for entry in ai_schedule:
                    if "slot" not in entry:
                        continue
                    slot_label = entry["slot"]
                    match = re.match(r'(\d{1,2}):00 (am|pm)', slot_label.split(' to ')[0])
                    if match:
                        hour = int(match.group(1))
                        if match.group(2) == 'pm' and hour != 12:
                            hour += 12
                        elif match.group(2) == 'am' and hour == 12:
                            hour = 0
                        hour_to_entry[hour] = entry
                complete_schedule = []
                missing = []
                for expected_slot in slot_labels:
                    match = re.match(r'(\d{1,2}):00 (am|pm)', expected_slot.split(' to ')[0])
                    if match:
                        hour = int(match.group(1))
                        if match.group(2) == 'pm' and hour != 12:
                            hour += 12
                        elif match.group(2) == 'am' and hour == 12:
                            hour = 0
                        if hour in hour_to_entry:
                            entry = hour_to_entry[hour]
                            entry["slot"] = expected_slot
                            complete_schedule.append(entry)
                        else:
                            missing.append(hour)
                if missing:
                    raise ValueError(f"Missing slots for hours: {missing}")
                parsed["schedule"] = complete_schedule
                data = parsed
                break

        except Exception as e:
            last_error = str(e)
            if attempt == 1:
                raise RuntimeError(f"AI generation failed after 2 attempts.\nLast error: {last_error}")
            else:
                continue

    if data is None:
        raise RuntimeError("Unexpected error: data is None")

    # Post-process: use AI's activity (short) and description
    schedule_dict = {entry["slot"]: entry for entry in data["schedule"]}
    final_schedule = []
    for slot_label in slot_labels:
        if slot_label == lunch_label:
            final_schedule.append({"slot": slot_label, "activity": "Lunch Break", "description": "Lunch Break"})
        else:
            short_key = full_to_short.get(slot_label, "")
            user_task = task_mapping.get(short_key, "")
            if slot_label in schedule_dict:
                ai_entry = schedule_dict[slot_label]
                if user_task == "-":
                    final_schedule.append({
                        "slot": slot_label,
                        "activity": "-",
                        "description": "-"
                    })
                else:
                    activity = ai_entry.get("activity", "")
                    description = ai_entry.get("description", "")
                    if not activity or activity == "No specific task":
                        words = user_task.split() if user_task else []
                        activity = " ".join(words[:5]) if words else "No specific task"
                    if not description or description == "No description provided.":
                        description = f"Worked on: {user_task}" if user_task else "No description provided."
                    final_schedule.append({
                        "slot": slot_label,
                        "activity": activity,
                        "description": description
                    })
            else:
                if user_task == "-":
                    final_schedule.append({"slot": slot_label, "activity": "-", "description": "-"})
                elif user_task:
                    words = user_task.split()
                    activity = " ".join(words[:5]) if words else "No specific task"
                    final_schedule.append({"slot": slot_label, "activity": activity, "description": f"Task: {user_task}"})
                else:
                    final_schedule.append({"slot": slot_label, "activity": "No specific task", "description": "No description provided."})

    data["schedule"] = final_schedule
    data["employee_name"] = data.get("employee_name", employee_name)
    data["position"] = data.get("position", position)
    data["date"] = data.get("date", report_date)
    data["_raw_response"] = raw_response
    return data

# ============= EXCEL GENERATION =============
def create_excel_from_schedule(schedule_data, template_bytes=None, time_slots=None):
    if template_bytes is None:
        template_bytes = DEFAULT_TEMPLATE_BYTES
    if time_slots is None:
        time_slots = get_time_slots(13)

    try:
        if template_bytes:
            wb = load_workbook(BytesIO(template_bytes))
            ws = wb.active
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws['B1'] = "EOD REPORT"
            ws.merge_cells('B1:C1')
            ws['B1'].font = Font(name='Calibri', size=16, bold=True)
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B3'] = "Name Of Employee"
            ws['C3'] = schedule_data.get("employee_name", DEFAULT_EMPLOYEE)
            ws['B4'] = "Position"
            ws['C4'] = schedule_data.get("position", DEFAULT_POSITION)
            ws['B5'] = "Date"
            date_val = schedule_data.get("date")
            if isinstance(date_val, str):
                try:
                    dt = parse_date(date_val)
                    ws['C5'] = dt.strftime("%d/%m/%Y")
                except:
                    ws['C5'] = date_val
            else:
                ws['C5'] = date_val.strftime("%d/%m/%Y") if isinstance(date_val, datetime) else date_val
            for r in [3,4,5]:
                ws[f'B{r}'].font = Font(bold=True)
            ws['B7'] = "Time"
            ws['C7'] = "Activity"
            ws['D7'] = "Description"
            for col in ['B','C','D']:
                ws[f'{col}7'].font = Font(bold=True)
            for i, slot in enumerate(time_slots):
                ws[f'B{8+i}'] = slot
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 28
            ws.column_dimensions['D'].width = 35
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = False
            thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows(min_row=7, max_row=7+len(time_slots), min_col=2, max_col=4):
                for cell in row:
                    cell.border = thin
            for cell_addr in ['C3','C4','C5']:
                ws[cell_addr].border = thin
    except Exception as e:
        st.error(f"Error creating workbook: {e}")
        raise

    def safe_write(cell_address, value):
        try:
            cell = ws[cell_address]
            for merged_range in ws.merged_cells:
                if cell.coordinate in merged_range:
                    top_left = merged_range.start_cell
                    top_left.value = value
                    return
            cell.value = value
        except:
            try:
                ws[cell_address] = value
            except:
                pass

    safe_write('C3', schedule_data.get("employee_name", DEFAULT_EMPLOYEE))
    safe_write('C4', schedule_data.get("position", DEFAULT_POSITION))
    date_val = schedule_data.get("date")
    if isinstance(date_val, str):
        try:
            dt = parse_date(date_val)
            safe_write('C5', dt.strftime("%d/%m/%Y"))
        except:
            safe_write('C5', date_val)
    else:
        safe_write('C5', date_val.strftime("%d/%m/%Y") if isinstance(date_val, datetime) else date_val)

    schedule_list = schedule_data.get("schedule", [])
    start_row = 8
    for idx, slot in enumerate(time_slots):
        row = start_row + idx
        entry = next((item for item in schedule_list if item.get("slot", "").strip().lower() == slot.lower()), None)
        activity = entry.get("activity", "No specific task") if entry else "No specific task"
        description = entry.get("description", "No description provided.") if entry else "No description provided."
        safe_write(f'C{row}', activity)
        safe_write(f'D{row}', description)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============= PDF GENERATION =============
def create_pdf_from_schedule(schedule_data, template_bytes=None, time_slots=None):
    if template_bytes is None:
        template_bytes = DEFAULT_TEMPLATE_BYTES
    excel_bytes = create_excel_from_schedule(schedule_data, template_bytes, time_slots)
    excel_data = excel_bytes.getvalue()

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_xlsx:
        tmp_xlsx.write(excel_data)
        xlsx_path = tmp_xlsx.name

    pdf_path = xlsx_path.replace('.xlsx', '.pdf')

    try:
        subprocess.run([
            'soffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', os.path.dirname(pdf_path),
            xlsx_path
        ], check=True, timeout=60)

        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()

        os.unlink(xlsx_path)
        os.unlink(pdf_path)

        return BytesIO(pdf_bytes)

    except (subprocess.CalledProcessError, FileNotFoundError, Exception) as e:
        if os.path.exists(xlsx_path):
            os.unlink(xlsx_path)
        if os.path.exists(pdf_path):
            os.unlink(pdf_path)
        st.warning("LibreOffice not available. Using fallback PDF.")
        return create_fallback_pdf(schedule_data, time_slots)

def create_fallback_pdf(schedule_data, time_slots=None):
    if time_slots is None:
        time_slots = get_time_slots(13)
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=15, leftMargin=15,
                            topMargin=15, bottomMargin=15)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('TitleStyle', parent=styles['Title'],
                                 alignment=1, fontSize=16, spaceAfter=8,
                                 fontName='Helvetica-Bold')
    elements.append(Paragraph("EOD REPORT", title_style))
    elements.append(Spacer(1, 4))

    detail_data = [
        ["Name Of Employee", schedule_data.get("employee_name", "N/A")],
        ["Position", schedule_data.get("position", "N/A")],
        ["Date", schedule_data.get("date", "N/A")]
    ]
    date_val = schedule_data.get("date")
    if isinstance(date_val, str):
        try:
            dt = parse_date(date_val)
            detail_data[2][1] = dt.strftime("%d/%m/%Y")
        except:
            detail_data[2][1] = date_val
    elif isinstance(date_val, datetime):
        detail_data[2][1] = date_val.strftime("%d/%m/%Y")

    detail_table = Table(detail_data, colWidths=[2.0*inch, 5.5*inch])
    detail_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('BOX', (0,0), (-1,-1), 0, colors.white),
        ('INNERGRID', (0,0), (-1,-1), 0, colors.white),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 10))

    schedule_list = schedule_data.get("schedule", [])
    table_data = [["Time", "Activity", "Description"]]
    for entry in schedule_list:
        table_data.append([
            entry.get("slot", ""),
            entry.get("activity", ""),
            entry.get("description", "")
        ])

    col_widths = [1.2*inch, 2.2*inch, 3.9*inch]
    schedule_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    schedule_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#CCCCCC")),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('WORDWRAP', (0,0), (-1,-1), True),
    ]))
    elements.append(schedule_table)

    doc.build(elements)
    pdf_data = buffer.getvalue()
    buffer.close()
    return BytesIO(pdf_data)

# ============= CONFETTI =============
def confetti():
    st.components.v1.html("""
    <script src="https://cdn.jsdelivr.net/npm/canvas-confetti@1"></script>
    <script>
        confetti({
            particleCount: 150,
            spread: 70,
            origin: { y: 0.6 }
        });
        setTimeout(() => {
            confetti({
                particleCount: 100,
                spread: 50,
                origin: { y: 0.5 }
            });
        }, 500);
    </script>
    """, height=0)

# ============= STREAMLIT UI =============
config = load_config()
saved_provider = config.get("provider", "Groq (Fastest)")
saved_model = config.get("model", "")
saved_api_key = config.get("api_key", "")

# ---- Sidebar ----
with st.sidebar:
    st.image("https://img.icons8.com/fluency/96/000000/google-forms.png", width=25)
    theme_label = "☀️ Light" if st.session_state.theme == "dark" else "🌙 Dark"
    if st.button(f"Switch to {theme_label} Theme", use_container_width=True):
        toggle_theme()

    # ---- Lunch Break Slider ----
    st.markdown("## 🕒 Lunch Break")
    if "prev_lunch_hour" not in st.session_state:
        st.session_state.prev_lunch_hour = 13

    lunch_hour = st.slider(
        "Select lunch break start:",
        min_value=12,
        max_value=14,
        value=st.session_state.prev_lunch_hour,
        step=1,
        format="%d:00 PM"
    )

    if lunch_hour != st.session_state.prev_lunch_hour:
        st.session_state.prev_lunch_hour = lunch_hour
        if "current_schedule" in st.session_state and st.session_state.current_schedule is not None:
            new_schedule = rebuild_schedule_for_lunch(lunch_hour, st.session_state.current_schedule)
            if new_schedule:
                st.session_state.current_schedule = new_schedule
                if "last_schedule" in st.session_state and st.session_state.last_schedule is not None:
                    old_schedule_data = st.session_state.last_schedule
                    schedule_data = {
                        "employee_name": old_schedule_data.get("employee_name", st.session_state.selected_employee_name),
                        "position": old_schedule_data.get("position", st.session_state.selected_employee_position),
                        "date": old_schedule_data.get("date", datetime.now().strftime("%d/%m/%Y")),
                        "schedule": new_schedule
                    }
                    st.session_state.last_schedule = schedule_data
                st.rerun()

    # ========== CONFIG SECTION ==========
    with st.expander("⚙️ Config", expanded=False):
        provider = st.selectbox(
            "AI Provider",
            options=list(PROVIDERS.keys()),
            index=list(PROVIDERS.keys()).index(saved_provider) if saved_provider in PROVIDERS else 0
        )
        model_name = st.text_input(
            "Model (optional)",
            placeholder=PROVIDERS[provider]["default_model"],
            value=saved_model if saved_provider == provider else ""
        )

        if PROVIDERS[provider]["api_key_required"]:
            if saved_api_key:
                st.success("✅ API key is set")
                if st.button("🔄 Change API Key"):
                    save_config(provider, model_name, "")
                    st.rerun()
            else:
                api_key = st.text_input("API Key", type="password")
                if st.button("💾 Save API Key"):
                    if api_key:
                        save_config(provider, model_name, api_key)
                        st.success("✅ API key saved")
                        st.rerun()
                    else:
                        st.warning("Please enter an API key")
                if "Groq" in provider:
                    st.caption("Get free key at [console.groq.com](https://console.groq.com)")
        else:
            api_key = None
            st.info("Ollama – no key needed.")

        if 'prev_provider' not in st.session_state:
            st.session_state.prev_provider = provider
            st.session_state.prev_model = model_name

        if (provider != st.session_state.prev_provider or
            model_name != st.session_state.prev_model):
            current_key = saved_api_key if saved_api_key else ""
            save_config(provider, model_name, current_key)
            st.session_state.prev_provider = provider
            st.session_state.prev_model = model_name
            st.success("✅ Config saved")

        if st.button("🗑️ Clear all config (including API key)"):
            clear_config()
            st.success("Config cleared.")
            st.rerun()

    with st.expander("👥 Employees", expanded=False):
        employees = load_employees()
        employee_names = [e["name"] for e in employees]
        if employee_names:
            selected_emp = st.selectbox("Select employee", employee_names,
                                        index=employee_names.index(st.session_state.selected_employee_name) if st.session_state.selected_employee_name in employee_names else 0)
            emp_details = next((e for e in employees if e["name"] == selected_emp), None)
            if emp_details:
                st.session_state.selected_employee_name = emp_details["name"]
                st.session_state.selected_employee_position = emp_details["position"]
        else:
            selected_emp = None

        with st.expander("➕ Add"):
            new_name = st.text_input("Name")
            new_pos = st.text_input("Position")
            if st.button("Add"):
                if new_name and new_pos:
                    add_employee(new_name, new_pos)
                    st.success(f"Added {new_name}")
                    st.rerun()
                else:
                    st.warning("Fill both")

        if selected_emp and st.button("🗑️ Delete"):
            delete_employee(selected_emp)
            st.success(f"Deleted {selected_emp}")
            st.rerun()

    st.divider()
    st.markdown("## 📁 Template")

    template_files = get_template_list()
    if DEFAULT_TEMPLATE_BYTES is not None:
        default_name = DEFAULT_TEMPLATE_FILE
        if default_name not in template_files:
            template_files.insert(0, default_name)
    template_options = ["Built-in"] + template_files

    selected_template = st.selectbox("Select Template", template_options)
    if selected_template == "Built-in":
        template_bytes = None
        st.info("Using built‑in layout")
    else:
        template_bytes = load_template_bytes(selected_template)
        if template_bytes is None:
            st.warning(f"Template '{selected_template}' not found. Using built‑in.")
            template_bytes = None
        else:
            st.success(f"Loaded '{selected_template}'")

    uploaded_file = st.file_uploader("Upload .xlsx", type=["xlsx"])
    if uploaded_file is not None:
        save_path = os.path.join(TEMPLATE_DIR, uploaded_file.name)
        with open(save_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        st.success(f"Saved '{uploaded_file.name}' to templates folder.")
        st.rerun()

    with st.expander("📜 History", expanded=False):
        if st.button("🗑️ Clear History", use_container_width=True):
            clear_history()
            st.success("History cleared.")
            st.rerun()

        all_history = load_history()
        employee_list = sorted(set(entry["employee_name"] for entry in all_history))
        selected_filter = st.selectbox("Filter", ["All"] + employee_list)
        if selected_filter != "All":
            filtered_history = [h for h in all_history if h["employee_name"] == selected_filter]
        else:
            filtered_history = all_history

        if filtered_history:
            groups = {}
            for entry in reversed(filtered_history):
                emp = entry.get("employee_name", "Unknown")
                if emp not in groups:
                    groups[emp] = []
                groups[emp].append(entry)
            for emp, entries in groups.items():
                with st.expander(f"👤 {emp} ({len(entries)})"):
                    for idx, entry in enumerate(entries[:5]):
                        display_date = entry.get("date", "N/A")
                        timestamp = entry.get("timestamp", str(idx))
                        if st.button(f"📂 {display_date}", key=f"load_{emp}_{timestamp}"):
                            st.session_state.loaded_report = entry
                            st.rerun()
                        st.caption(f"Position: {entry.get('position', '')}")
        else:
            st.write("No reports yet.")

        if "loaded_report" in st.session_state and st.session_state.loaded_report is not None:
            if st.button("🗑️ Clear loaded", use_container_width=True):
                st.session_state.loaded_report = None
                st.session_state.current_schedule = None
                st.session_state.loaded_date = None
                st.rerun()

# ---- Load history if present ----
if st.session_state.get("loaded_report") is not None:
    report = st.session_state.loaded_report
    # Update employee, position, date
    st.session_state.selected_employee_name = report["employee_name"]
    st.session_state.selected_employee_position = report["position"]
    # Parse date and store for date input
    try:
        st.session_state.loaded_date = parse_date(report["date"])
    except:
        st.session_state.loaded_date = datetime.now()
    # Set the current schedule to the loaded one
    st.session_state.current_schedule = report["schedule"]
    # Also set last_schedule so that download buttons work
    st.session_state.last_schedule = {
        "employee_name": report["employee_name"],
        "position": report["position"],
        "date": report["date"],
        "schedule": report["schedule"]
    }
    # We keep loaded_report so that we can show a "Clear loaded" button.

# ---- Session state ----
if "loaded_report" not in st.session_state:
    st.session_state.loaded_report = None
if "last_schedule" not in st.session_state:
    st.session_state.last_schedule = None
if "last_input" not in st.session_state:
    st.session_state.last_input = ""
if "last_config" not in st.session_state:
    st.session_state.last_config = {}
if "current_schedule" not in st.session_state:
    st.session_state.current_schedule = None
if "raw_response" not in st.session_state:
    st.session_state.raw_response = None
if "loaded_date" not in st.session_state:
    st.session_state.loaded_date = None

# ---- Main area ----
left_col, right_col = st.columns([0.4, 0.6], gap="small")

with left_col:
    st.markdown("### ✍️ Inputs")
    employees = load_employees()
    employee_names = [e["name"] for e in employees]
    if employee_names:
        current_name = st.session_state.selected_employee_name
        if current_name not in employee_names:
            current_name = employee_names[0]
        selected_emp_main = st.selectbox("👤 Employee", employee_names, index=employee_names.index(current_name))
        emp_details_main = next((e for e in employees if e["name"] == selected_emp_main), None)
        if emp_details_main:
            st.session_state.selected_employee_name = emp_details_main["name"]
            st.session_state.selected_employee_position = emp_details_main["position"]
    else:
        st.session_state.selected_employee_name = DEFAULT_EMPLOYEE
        st.session_state.selected_employee_position = DEFAULT_POSITION

    position = st.text_input("💼 Position", value=st.session_state.selected_employee_position)
    # Use loaded date if available, else today
    default_date = st.session_state.loaded_date if st.session_state.loaded_date is not None else datetime.now()
    report_date = st.date_input("📅 Date", value=default_date, format="DD/MM/YYYY")

    st.markdown("### 📝 Task per Time Slot")
    st.caption("Type your task after each time. Use '-' to indicate nothing was done.")

    time_slots_full = get_time_slots(lunch_hour)
    time_slots_short = get_time_slots_short(lunch_hour)
    lunch_index = lunch_hour - 10

    if "slot_tasks" not in st.session_state:
        st.session_state.slot_tasks = {slot: "" for slot in time_slots_full if slot != time_slots_full[lunch_index]}

    for i, full_slot in enumerate(time_slots_full):
        if i == lunch_index:
            st.markdown(f"""
            <div class="slot-card lunch-card" style="display: flex; align-items: center; padding: 0.2rem 0.5rem;">
                <span class="slot-label">🍴 {full_slot}</span>
                <span style="flex:1; text-align:center; color: #ffc107;">Lunch Break</span>
            </div>
            """, unsafe_allow_html=True)
            continue

        short_time = time_slots_short[i]
        current_val = st.session_state.slot_tasks.get(full_slot, "")
        col_label, col_input = st.columns([0.3, 0.7])
        with col_label:
            st.markdown(f"<span class='slot-label'>{short_time}</span>", unsafe_allow_html=True)
        with col_input:
            new_val = st.text_input(
                label="",
                value=current_val,
                key=f"slot_{i}",
                placeholder="e.g., posted stories",
                label_visibility="collapsed"
            )
            st.session_state.slot_tasks[full_slot] = new_val

    task_lines = []
    for full_slot in time_slots_full:
        if full_slot == time_slots_full[lunch_index]:
            continue
        short_time = time_slots_short[time_slots_full.index(full_slot)]
        task = st.session_state.slot_tasks.get(full_slot, "").strip()
        if task:
            task_lines.append(f"{short_time}: {task}")
        else:
            task_lines.append(f"{short_time}: ")
    task_summary = "\n".join(task_lines)
    st.session_state.task_summary = task_summary

    templates_quick = {
        "Feng Shui & Content": {
            "10:00-11:00": "Posted Feng Shui stories",
            "11:00-12:00": "Created post on July animal signs",
            "12:00-13:00": "Started editing Kedarnath reel",
            "14:00-15:00": "Created 15 AI creatives for reel",
            "15:00-16:00": "Continued editing reel",
            "16:00-17:00": "Reviewed performance",
            "17:00-18:00": "Planned next steps"
        },
        "Meetings & Documentation": {
            "10:00-11:00": "Team sync meeting",
            "11:00-12:00": "Wrote meeting notes",
            "12:00-13:00": "Follow-up emails",
            "14:00-15:00": "Project planning",
            "15:00-16:00": "Client call",
            "16:00-17:00": "Prepared status report",
            "17:00-18:00": "Reviewed and finalized"
        },
        "Development & Testing": {
            "10:00-11:00": "Fixed bugs",
            "11:00-12:00": "Developed new feature",
            "12:00-13:00": "Code review",
            "14:00-15:00": "Wrote tests",
            "15:00-16:00": "Deployed to staging",
            "16:00-17:00": "Updated documentation",
            "17:00-18:00": "Sprint planning"
        },
        "Custom": {}
    }
    selected_template_quick = st.selectbox("📝 Quick template", list(templates_quick.keys()), key="quick_template")
    if selected_template_quick != "Custom" and templates_quick[selected_template_quick]:
        if st.button("📋 Load template", use_container_width=True):
            for short, task in templates_quick[selected_template_quick].items():
                for full_slot in time_slots_full:
                    if time_slots_short[time_slots_full.index(full_slot)] == short:
                        st.session_state.slot_tasks[full_slot] = task
                        break
            st.rerun()

    col_gen, col_reg = st.columns(2)
    with col_gen:
        generate_clicked = st.button("🚀 Generate", type="primary", use_container_width=True)
    with col_reg:
        regenerate_clicked = st.button("🔄 Regenerate", use_container_width=True,
                                       disabled=st.session_state.last_schedule is None)

with right_col:
    if st.session_state.current_schedule is not None:
        st.markdown("### 📊 Edit Schedule")
        st.caption("Edit Activity and Description directly below.")

        schedule_list = st.session_state.current_schedule
        edited_schedule = []

        for idx, entry in enumerate(schedule_list):
            cols = st.columns([2, 2, 3])
            with cols[0]:
                st.markdown(f"**{entry['slot']}**")
            with cols[1]:
                new_activity = st.text_input(
                    "Activity",
                    value=entry["activity"],
                    key=f"edit_act_{idx}",
                    label_visibility="collapsed"
                )
            with cols[2]:
                new_description = st.text_input(
                    "Description",
                    value=entry["description"],
                    key=f"edit_desc_{idx}",
                    label_visibility="collapsed"
                )
            edited_schedule.append({
                "slot": entry["slot"],
                "activity": new_activity,
                "description": new_description
            })

        st.session_state.current_schedule = edited_schedule

        schedule_data = {
            "employee_name": st.session_state.selected_employee_name,
            "position": st.session_state.selected_employee_position,
            "date": report_date.strftime("%d/%m/%Y"),
            "schedule": edited_schedule
        }
        st.session_state.last_schedule = schedule_data

        time_slots = get_time_slots(lunch_hour)
        excel_data = create_excel_from_schedule(schedule_data, template_bytes, time_slots)
        pdf_data = create_pdf_from_schedule(schedule_data, template_bytes, time_slots)

        file_date = parse_date(schedule_data["date"])
        emp_name = schedule_data["employee_name"]
        excel_filename = generate_filename(file_date, emp_name, "xlsx")
        pdf_filename = generate_filename(file_date, emp_name, "pdf")

        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📥 Download Excel",
                data=excel_data,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col2:
            st.download_button(
                label="📄 Download PDF",
                data=pdf_data,
                file_name=pdf_filename,
                mime="application/pdf",
                use_container_width=True
            )
    else:
        st.info("👈 Fill in your per‑slot tasks and click Generate.")

# ---- Generation logic ----
if generate_clicked or regenerate_clicked:
    if regenerate_clicked:
        tasks = st.session_state.last_input
        cfg = st.session_state.last_config
        provider_used = cfg["provider"]
        api_key_used = cfg["api_key"]
        model_used = cfg["model"]
        emp_used = cfg["employee"]
        pos_used = cfg["position"]
        date_used = cfg["date"]
        lunch_hour_used = cfg.get("lunch_hour", 13)
    else:
        tasks = st.session_state.task_summary.strip()
        if not tasks:
            st.warning("Please describe your daily tasks (at least one slot).")
            st.stop()
        
        if PROVIDERS[provider]["api_key_required"]:
            saved_key = load_config().get("api_key", "")
            if saved_key:
                api_key = saved_key
            else:
                api_key = ""
        else:
            api_key = None
            
        emp_used = st.session_state.selected_employee_name
        pos_used = st.session_state.selected_employee_position
        date_used = report_date.strftime("%d/%m/%Y")
        lunch_hour_used = lunch_hour
        st.session_state.last_input = tasks
        st.session_state.last_config = {
            "provider": provider,
            "api_key": api_key,
            "model": model_name,
            "employee": emp_used,
            "position": pos_used,
            "date": date_used,
            "lunch_hour": lunch_hour_used
        }
        provider_used = provider
        api_key_used = api_key
        model_used = model_name

    progress_bar = st.progress(0)
    status_text = st.empty()
    
    def update_progress(progress, message):
        progress_bar.progress(progress)
        status_text.markdown(f"<span style='color:#667eea; font-weight:500;'>{message}</span>", unsafe_allow_html=True)

    try:
        update_progress(10, "🚀 Initializing...")
        data = generate_schedule(tasks, emp_used, pos_used, date_used, provider_used, api_key_used, model_used, lunch_hour_used, progress_callback=update_progress)
        update_progress(95, "✨ Finalizing...")

        st.session_state.raw_response = data.pop("_raw_response", None)

        st.session_state.last_schedule = data
        add_history_entry(emp_used, pos_used, date_used, data["schedule"])
        emp_list = load_employees()
        if not any(e["name"] == emp_used for e in emp_list):
            add_employee(emp_used, pos_used)

        st.session_state.current_schedule = data["schedule"]

        st.balloons()
        confetti()
        st.success("✅ Report generated successfully! 🎉")
        status_text.empty()
        progress_bar.empty()
        st.rerun()
    except Exception as e:
        st.error(f"❌ {e}")
        progress_bar.empty()
        status_text.empty()
        st.session_state.last_schedule = None